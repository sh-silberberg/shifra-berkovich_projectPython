from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
import os
import openpyxl
from openpyxl import load_workbook
import pandas as pd

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'No file part'
    
    file = request.files['file']

    if file.filename == '':
        return 'No selected file'

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
    file.save(file_path)

    try:
        workbook = load_workbook(filename=file_path)
        number_of_sheets = len(workbook.sheetnames)
        return jsonify({
            "file_url": file_path,
            "number_of_sheets": number_of_sheets
        })
    except Exception as e:
        return f"An error occurred: {e}"

@app.route('/generate_report', methods=['POST'])
def generate_report():
    data = request.json
    file_path = data.get('file_path')
    sheets = data.get('sheets')
    report_data = {
        'report': 'Sample report data',  
    }

    return jsonify(report_data)

@app.route('/generate_pdf_report', methods=['POST'])
def generate_pdf_report():
    report_data = request.json
    # PDF generation logic here
    return send_file('path_to_generated_pdf_report.pdf', as_attachment=True)

if __name__ == '__main__':
    app.run()




def get_number_of_sheets(file_path):
    workbook = openpyxl.load_workbook(file_path)
    return len(workbook.sheetnames)


file_path = 'E:/python/2/try.xlsx'
print(get_number_of_sheets(file_path))


def calculate_total_sum(file_path):
    workbook = openpyxl.load_workbook(file_path)
    total_sum = 0
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    total_sum += cell

    return total_sum

print(calculate_total_sum(file_path))
import matplotlib.pyplot as plt



def display_sheet_sums_bar_chart(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet_sums = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        total_sum = 0
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    total_sum += cell

        sheet_sums[sheet_name] = total_sum

    plt.bar(sheet_sums.keys(), sheet_sums.values())
    plt.xlabel('Sheet Name')
    plt.ylabel('Total Sum')
    plt.title('Total Sum of Each Sheet')
    plt.show()

display_sheet_sums_bar_chart(file_path)







def calculate_average_across_files(file_paths):
    total_sum = 0
    total_count = 0

    for file_path in file_paths:
        df = pd.read_excel(file_path)
        
        # Calculate sum and count for the current file
        file_sum = df.stack().sum()
        file_count = df.stack().count()
        
        total_sum += file_sum
        total_count += file_count

        print(f"Total sum for {file_path}: {file_sum}")
        print(f"Total count for {file_path}: {file_count}")

    if total_count == 0:
        return 0
    
    
    
    return total_sum / total_count

    print("Total sum across all files:", total_sum)
    print("Total count across all files:", total_count)
    
file_paths = ['E:/python/2/try.xlsx']  # Update with full file paths
average = calculate_average_across_files(file_paths)
print("Average across files:", average)


             
             
             
from reportlab.pdfgen import canvas

# Function to create the PDF report
def create_report(file_path):
    df = pd.read_excel(file_path)
    
    # Logic to calculate total sum and averages
    total_sum = df.stack().sum()
    averages = df.mean(axis=1)
    
    # Logic to create graphs (you can customize this as needed)
    plt.figure(figsize=(5, 5))
    plt.bar(df.columns, df.sum())
    plt.title('pdf per sheet')
    plt.savefig('PDF.png')
    
    # Add more graph logic here for additional graphs
    
    # Create the PDF report
    c = canvas.Canvas('E:/python/2/try.xlsx')  
    c.drawString(400, 1000, "Report for file: {}".format(file_path))
    c.drawString(400, 880, "Total Sum: {}".format(total_sum))
    
    # Write data to PDF
    y_position = 760
    for sheet_name, avg in averages.items():
        c.drawString(100, y_position, "Sheet Name: {}, Average: {}".format(sheet_name, avg))
        y_position -= 20
    
    # Add graphs to the PDF
    c.drawImage('E:/python/2/try.xlsx', 4, 0.5)  # Add image to PDF
    c.showPage()
    
    # Save the PDF
    c.save()

# Example usage
file_path = 'E:/python/2/try.xlsx'  # Update with your file path
create_report(file_path)


